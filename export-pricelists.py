"""
reCarpet B2B — SparkLayer Price List Exporter
----------------------------------------------
Run this script after updating base (brutto) prices in reCarpet-pricelists-master.xlsx
('Members' sheet). It generates 16 CSV files — 4 customer tiers × 4 currencies.

Tier structure:
    utloggad  — guest pricing (brutto, no discount)
    member    — Member (brutto, no discount)  [Shopify tag: member]
    plus      — Member Plus (10 % discount)   [Shopify tags: member, plus]
    premium   — Member Premium (10 % discount, + services access)
                                               [Shopify tags: member, premium]
                Krets customers also use this price list
                                               [Shopify tags: member, premium, krets]

Usage:
    python export-pricelists.py

Output: sparklayer-pricelists/ folder (16 CSV files)
"""

import os
import csv
import math
from openpyxl import load_workbook

EXCEL_FILE = "reCarpet-pricelists-master.xlsx"
OUTPUT_DIR = "sparklayer-pricelists"
SOURCE_SHEET = "Members"  # brutto / source of truth
DATA_START_ROW = 5

# Tier name → discount multiplier applied to brutto SEK price
TIERS = {
    "utloggad": 1.00,
    "member":   1.00,
    "plus":     0.90,
    "premium":  0.90,
}

# Currency → row in Exchange Rates sheet (column C holds the rate)
CURRENCIES = {
    "sek": None,   # base — rate resolved at runtime (= 1.0)
    "nok": 4,
    "dkk": 5,
    "eur": 6,
}


def round_up(value, decimals=2):
    """Round ALWAYS upward to N decimals (confirmed by Hampus)."""
    factor = 10 ** decimals
    return math.ceil(value * factor) / factor


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)
    output_dir = os.path.join(script_dir, OUTPUT_DIR)
    os.makedirs(output_dir, exist_ok=True)

    print(f"Reading: {EXCEL_FILE}")
    wb = load_workbook(excel_path, data_only=True)

    ws_rates = wb["Exchange Rates"]
    rates = {"sek": 1.0}
    for cur, row in CURRENCIES.items():
        if row is None:
            continue
        rates[cur] = ws_rates.cell(row=row, column=3).value
    print(f"  FX — NOK: {rates['nok']}, DKK: {rates['dkk']}, EUR: {rates['eur']}")

    ws = wb[SOURCE_SHEET]
    base_rows = []
    row = DATA_START_ROW
    while True:
        sku = ws.cell(row=row, column=1).value
        if sku is None:
            break
        qty = ws.cell(row=row, column=2).value
        sek = ws.cell(row=row, column=3).value
        if sek is None:
            print(f"  ! skipping {sku}: no SEK price")
            row += 1
            continue
        base_rows.append((sku, qty or 1, float(sek)))
        row += 1
    print(f"  Base SKUs loaded: {len(base_rows)}")

    total_files = 0
    for tier_name, tier_mult in TIERS.items():
        for currency, rate in rates.items():
            csv_name = f"sparklayer-{tier_name}-{currency}.csv"
            csv_path = os.path.join(output_dir, csv_name)

            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["sku", "quantity", "price"])
                for sku, qty, sek in base_rows:
                    price = round_up(sek * tier_mult * rate, 2)
                    writer.writerow([sku, qty, f"{price:.2f}"])

            print(f"  ✓ {csv_name}  ({len(base_rows)} rows)  "
                  f"[tier={tier_name} discount={(1-tier_mult)*100:.0f}% fx={rate}]")
            total_files += 1

    print(f"\nDone — {total_files} files saved to '{OUTPUT_DIR}/'")
    print("Upload each file to the matching SparkLayer price list:")
    print("  utloggad-*  → SparkLayer 'Guest' price list (non-logged-in pricing)")
    print("  member-*    → SparkLayer 'Member' price list   (tag: member)")
    print("  plus-*      → SparkLayer 'Plus' price list     (tag: member, plus)")
    print("  premium-*   → SparkLayer 'Premium' price list  (tag: member, premium)")
    print("                (Krets customers use this list via tag: member, premium, krets)")


if __name__ == "__main__":
    main()
