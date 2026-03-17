"""
reCarpet — Orak Import Script (New CSV Format)
===============================================
Läser ORAKs standard CSV-export och:
  1. Skapar/uppdaterar produkter i Shopify via Admin API
  2. Sätter meta fields (dimensioner, teknisk spec, etc.)
  3. Genererar SparkLayer-prislistor för Silver (0%) och Guld (-10%)
  4. Publicerar produkter till SparkLayer B2B-kanal automatiskt
  5. Lägger till produkter i rätt produktserie (collection)

Usage:
    python import-orak.py --csv path/to/orak.csv
    python import-orak.py --csv path/to/orak.csv --dry-run
    python import-orak.py --csv path/to/orak.csv --pricelists-only

Krav:
    pip install pandas requests python-dotenv

Setup:
    1. Kopiera .env.example → .env och fyll i credentials
    2. Kör setup-metafields.py en gång för att skapa meta field definitions
    3. Kör detta script med ORAKs CSV-fil
"""

import os
import sys
import csv
import json
import time
import argparse
import pandas as pd
import requests
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

# ─── Konfiguration ────────────────────────────────────────────────────────────

SHOPIFY_SHOP  = os.getenv("SHOPIFY_SHOP")    # t.ex. recarpet.myshopify.com
SHOPIFY_TOKEN = os.getenv("SHOPIFY_TOKEN")   # Admin API access token
API_VERSION   = "2024-01"

BASE_DIR   = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
SYNC_LOG   = OUTPUT_DIR / "sync-log.json"

SPARKLAYER_DIR = BASE_DIR.parent / "sparklayer-pricelists"

# SparkLayer pricelist multipliers
# Brons (public): no pricelist — sees standard Shopify price
# Silver (member): netto x 1.00 (0% discount)
# Guld (entrepreneur): netto x 0.90 (10% discount)
# Krets: netto x 0.80 (20% discount off entrepreneur price)
SILVER_MULTIPLIER = 1.00
GULD_MULTIPLIER   = 0.90
KRETS_MULTIPLIER  = 0.80

# Fallback-valutakurser (används om live-hämtning misslyckas)
FALLBACK_RATES = {
    "EUR_SEK": 11.20,
    "EUR_NOK": 11.70,
    "EUR_DKK": 7.46,
    "source":  "fallback",
}

# Sökväg till Excel-masterfile för prislistor (relativt till detta script)
PRICELIST_EXCEL = BASE_DIR.parent / "reCarpet-pricelists-master.xlsx"

# SparkLayer sales channel name (exact title as it appears in Shopify Admin → Settings → Sales channels)
SPARKLAYER_CHANNEL_NAME = "SparkLayer B2B & Wholesale"

# Produktserie — alla ORAK-produkter hamnar i "Återbrukade mattor"
# OBS: handlen ändras INTE när du byter namn i Shopify Admin.
# Kolla rätt handle: Admin → Collections → klicka samlingen → rulla ner till "Sökmotor"-sektionen
ATERVUNNA_HANDLE = "atervunna-mattor"  # ändra om handlen är annorlunda, t.ex. "mattor"

# "Produit à venir" = återbrukade mattor under rengöring → publiceras direkt men märks med tagg "kommande"
KOMMANDE_LABELS = {"produit a venir", "a venir"}

# Rader som ska hoppas över — ORAKs interna anteckningar, ej produkter
SKIP_RAW_SKUS = {
    "Mr_Rouni",
    "LAB SERVICES",
    "Yannick_16/12_sur chantier",
}
SKIP_PATTERNS = ["bloqué", "bloque", "sur chantier"]

# ORAK CSV column → internal key mapping
COLUMN_MAP = {
    "title":             "title",
    "brand":             "brand",
    "productType":       "product_label",
    "identificationID":  "sku",
    "dimensions":        "dimensions",
    "pricePerUnit":      "price",
    "quantityAvailable": "quantity",
    "mainImage":         "image_url",
    "technicalSheetURL": "technical_sheet_url",
}


# ─── Shopify API ──────────────────────────────────────────────────────────────

def api_headers():
    return {
        "X-Shopify-Access-Token": SHOPIFY_TOKEN,
        "Content-Type": "application/json",
    }


def shopify_get(endpoint, params=None):
    url = f"https://{SHOPIFY_SHOP}/admin/api/{API_VERSION}/{endpoint}"
    r = requests.get(url, headers=api_headers(), params=params)
    _rate_limit(r)
    r.raise_for_status()
    return r.json()


def shopify_post(endpoint, payload):
    url = f"https://{SHOPIFY_SHOP}/admin/api/{API_VERSION}/{endpoint}"
    r = requests.post(url, headers=api_headers(), json=payload)
    _rate_limit(r)
    r.raise_for_status()
    return r.json()


def shopify_put(endpoint, payload):
    url = f"https://{SHOPIFY_SHOP}/admin/api/{API_VERSION}/{endpoint}"
    r = requests.put(url, headers=api_headers(), json=payload)
    _rate_limit(r)
    r.raise_for_status()
    return r.json()


def _rate_limit(response):
    if response.status_code == 429:
        wait = int(response.headers.get("Retry-After", 2))
        print(f"    Waiting {wait}s (rate limit)...")
        time.sleep(wait)
    else:
        header = response.headers.get("X-Shopify-Shop-Api-Call-Limit", "")
        if header:
            used, total = map(int, header.split("/"))
            if used / total > 0.8:
                time.sleep(0.5)
        else:
            time.sleep(0.25)


def graphql(query: str, variables: dict = None) -> dict:
    """Kör en GraphQL-query mot Shopify Admin API."""
    url = f"https://{SHOPIFY_SHOP}/admin/api/{API_VERSION}/graphql.json"
    payload = {"query": query}
    if variables:
        payload["variables"] = variables
    r = requests.post(url, headers=api_headers(), json=payload)
    _rate_limit(r)
    r.raise_for_status()
    return r.json()


def get_sparklayer_publication_gid() -> str | None:
    """
    Hämtar SparkLayer-kanalens GraphQL-ID via Admin API GraphQL.
    REST publications.json togs bort i Shopify API 2022-07 — använder GraphQL istället.
    """
    query = """
    query {
      publications(first: 20) {
        edges {
          node {
            id
            name
          }
        }
      }
    }
    """
    try:
        result = graphql(query)
        # Logga GraphQL-fel (t.ex. saknat scope) för enkel felsökning
        gql_errors = result.get("errors")
        if gql_errors:
            msgs = [e.get("message", "") for e in gql_errors]
            print(f"  Warning: GraphQL-fel vid hämtning av publications: {msgs}")
            print(f"  Tips: Kontrollera att din Shopify-app har scopes: read_publications, write_publications")
            return None
        data = result.get("data") or {}
        edges = (data.get("publications") or {}).get("edges", [])
        for edge in edges:
            node = edge["node"]
            if SPARKLAYER_CHANNEL_NAME.lower() in node.get("name", "").lower():
                print(f"  SparkLayer GID: {node['id']}  ({node['name']})")
                return node["id"]
        names = [e["node"]["name"] for e in edges]
        print(f"  Warning: SparkLayer channel ej hittad. Tillgängliga: {names}")
    except Exception as e:
        print(f"  Warning: kunde inte hämta publications via GraphQL: {e}")
    return None


def publish_to_sparklayer(product_id: int, publication_gid: str):
    """Publicerar en produkt till SparkLayer via GraphQL publishablePublish."""
    mutation = """
    mutation publishablePublish($id: ID!, $input: [PublicationInput!]!) {
      publishablePublish(id: $id, input: $input) {
        userErrors { field message }
      }
    }
    """
    product_gid = f"gid://shopify/Product/{product_id}"
    try:
        result = graphql(mutation, {
            "id": product_gid,
            "input": [{"publicationId": publication_gid}],
        })
        data = result.get("data") or {}
        errors = (data.get("publishablePublish") or {}).get("userErrors", [])
        if errors:
            print(f"      Warning: SparkLayer-publicering userErrors: {errors}")
    except Exception as e:
        print(f"      Warning: SparkLayer-publicering misslyckades för {product_id}: {e}")


def get_collection_id_map() -> dict:
    """
    Returnerar {handle: collection_id} och {title_lower: collection_id}
    för alla custom collections — söker på både handle och titel.
    """
    id_map = {}
    try:
        data = shopify_get("custom_collections.json", {"limit": 250, "fields": "id,handle,title"})
        for col in data.get("custom_collections", []):
            id_map[col["handle"]] = col["id"]
            id_map[col["title"].lower()] = col["id"]   # fallback på titel
    except Exception as e:
        print(f"  Warning: kunde inte hämta collections: {e}")
    return id_map


def add_product_to_collection(product_id: int, collection_id: int):
    """Lägger till en produkt i en collection (skapar collect om den inte finns)."""
    try:
        shopify_post("collects.json", {
            "collect": {"product_id": product_id, "collection_id": collection_id}
        })
    except Exception as e:
        # 422 = redan i collection, ignorera
        if "422" not in str(e):
            print(f"      Warning: collection-tilldelning misslyckades: {e}")


def get_existing_skus() -> dict:
    print("  Fetching existing products from Shopify...")
    sku_map = {}
    params = {"limit": 250, "fields": "id,title,variants"}
    data = shopify_get("products.json", params)
    for product in data.get("products", []):
        for variant in product.get("variants", []):
            if variant.get("sku"):
                sku_map[variant["sku"]] = {
                    "product_id": product["id"],
                    "variant_id": variant["id"],
                    "inventory_item_id": variant.get("inventory_item_id"),
                }
    print(f"    {len(sku_map)} existing products found")
    return sku_map


def get_first_location_id() -> int:
    locations = shopify_get("locations.json")
    return locations["locations"][0]["id"]


def create_product(title, vendor, product_type, tags, price, sku, image_url) -> dict:
    payload = {
        "product": {
            "title": title,
            "vendor": vendor,
            "product_type": product_type,
            "tags": tags,
            "published": True,
            "variants": [{
                "sku": sku,
                "price": f"{price:.2f}",
                "inventory_management": "shopify",
                "requires_shipping": False,
                "taxable": True,
            }],
        }
    }
    if image_url:
        payload["product"]["images"] = [{"src": image_url, "alt": title}]
    return shopify_post("products.json", payload)["product"]


def update_product(product_id, variant_id, title, vendor, product_type, tags, price, image_url):
    """Uppdaterar alla produktfält: titel, varumärke, typ, taggar, pris och bild."""
    payload = {
        "product": {
            "id": product_id,
            "title": title,
            "vendor": vendor,
            "product_type": product_type,
            "tags": tags,
            "variants": [{"id": variant_id, "price": f"{price:.2f}"}],
        }
    }
    if image_url:
        payload["product"]["images"] = [{"src": image_url, "alt": title}]
    shopify_put(f"products/{product_id}.json", payload)


def update_variant_price(variant_id, price):
    shopify_put(f"variants/{variant_id}.json", {
        "variant": {"id": variant_id, "price": f"{price:.2f}"}
    })


def set_inventory(inventory_item_id, location_id, quantity):
    shopify_post("inventory_levels/set.json", {
        "location_id": location_id,
        "inventory_item_id": inventory_item_id,
        "available": int(quantity),
    })


def set_metafields(product_id, metafields: list):
    for mf in metafields:
        if not mf.get("value"):
            continue
        try:
            shopify_post(f"products/{product_id}/metafields.json", {"metafield": mf})
        except Exception as e:
            print(f"      Warning: metafield '{mf['key']}' failed: {e}")


# ─── Data loading ─────────────────────────────────────────────────────────────

def load_orak_csv(path: Path) -> list:
    df = pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
    df = df.rename(columns=COLUMN_MAP)

    def fix_url(x):
        if isinstance(x, str) and x.startswith("//"):
            return "https:" + x
        return x

    if "image_url" in df.columns:
        df["image_url"] = df["image_url"].apply(fix_url)
    if "technical_sheet_url" in df.columns:
        df["technical_sheet_url"] = df["technical_sheet_url"].apply(fix_url)

    df = df[df["sku"].notna() & (df["sku"].astype(str).str.strip() != "")]
    df = df[df["title"].notna() & (df["title"].astype(str).str.strip() != "")]
    return df.to_dict("records")


def safe(val, fallback="") -> str:
    if pd.isna(val) if hasattr(pd, 'isna') else val != val:
        return fallback
    s = str(val).strip()
    return fallback if s in {"nan", "None", ""} else s


# ─── Sync log ─────────────────────────────────────────────────────────────────

def load_sync_log() -> dict:
    if SYNC_LOG.exists():
        with open(SYNC_LOG) as f:
            return json.load(f)
    return {}


def save_sync_log(log: dict):
    OUTPUT_DIR.mkdir(exist_ok=True)
    with open(SYNC_LOG, "w") as f:
        json.dump(log, f, indent=2)


# ─── SparkLayer pricelists ─────────────────────────────────────────────────────

def fetch_exchange_rates() -> dict:
    """
    Hämtar live valutakurser från frankfurter.app (gratis, ingen API-nyckel krävs).
    Returnerar EUR→SEK, EUR→NOK, EUR→DKK.
    Använder FALLBACK_RATES om hämtningen misslyckas.
    """
    try:
        r = requests.get(
            "https://api.frankfurter.app/latest?from=EUR&to=SEK,NOK,DKK",
            timeout=5
        )
        r.raise_for_status()
        data = r.json()["rates"]
        rates = {
            "EUR_SEK": round(data["SEK"], 4),
            "EUR_NOK": round(data["NOK"], 4),
            "EUR_DKK": round(data["DKK"], 4),
            "source":  "live",
            "date":    r.json().get("date", "okänt"),
        }
        print(f"  Valutakurser ({rates['date']}): 1 EUR = {rates['EUR_SEK']} SEK | {rates['EUR_NOK']} NOK | {rates['EUR_DKK']} DKK")
        return rates
    except Exception as e:
        print(f"  Warning: live-kurser ej tillgängliga ({e}). Använder fallback: {FALLBACK_RATES}")
        return FALLBACK_RATES.copy()


def generate_pricelists(products: list, rates: dict):
    """
    Genererar alla 8 SparkLayer-prislistor (2 kundnivåer × 4 valutor).
    ORAKs priser är i EUR — konverteras till SEK/NOK/DKK via rates.
    Format: sku, quantity, price  (quantity=1 = gäller alla kvantiteter)
    """
    SPARKLAYER_DIR.mkdir(exist_ok=True)

    eur_sek = rates["EUR_SEK"]
    eur_nok = rates["EUR_NOK"]
    eur_dkk = rates["EUR_DKK"]

    # Prislistor: {filnamn: (multiplier, kurs)}
    LISTS = {
        "sparklayer-member-sek.csv":       (SILVER_MULTIPLIER, eur_sek),
        "sparklayer-member-nok.csv":       (SILVER_MULTIPLIER, eur_nok),
        "sparklayer-member-dkk.csv":       (SILVER_MULTIPLIER, eur_dkk),
        "sparklayer-member-eur.csv":       (SILVER_MULTIPLIER, 1.0),
        "sparklayer-entrepreneur-sek.csv": (GULD_MULTIPLIER,   eur_sek),
        "sparklayer-entrepreneur-nok.csv": (GULD_MULTIPLIER,   eur_nok),
        "sparklayer-entrepreneur-dkk.csv": (GULD_MULTIPLIER,   eur_dkk),
        "sparklayer-entrepreneur-eur.csv": (GULD_MULTIPLIER,   1.0),
        "sparklayer-krets-sek.csv":        (KRETS_MULTIPLIER,  eur_sek),
    }

    # Bygg rader per lista
    list_rows = {name: [] for name in LISTS}

    for p in products:
        raw_sku = safe(p.get("sku"))
        if not raw_sku or should_skip(raw_sku):
            continue
        sku = make_recarpet_sku(raw_sku)
        try:
            eur_price = float(p.get("price", 0) or 0)
        except (ValueError, TypeError):
            continue
        if eur_price <= 0:
            continue

        for name, (mult, rate) in LISTS.items():
            price = round(eur_price * rate * mult, 2)
            list_rows[name].append([sku, 1, f"{price:.2f}"])

    def write_csv(filename, rows):
        path = SPARKLAYER_DIR / filename
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["sku", "quantity", "price"])
            writer.writerows(rows)
        print(f"  OK  {filename}  ({len(rows)} SKUs)")

    print("\n  Genererar SparkLayer-prislistor (9 st)...")
    for name, rows in list_rows.items():
        write_csv(name, rows)

    rate_src = rates.get("source", "?")
    rate_date = rates.get("date", "")
    print(f"""
  Kurs ({rate_src}{' ' + rate_date if rate_date else ''}):
    1 EUR = {eur_sek:.4f} SEK | {eur_nok:.4f} NOK | {eur_dkk:.4f} DKK

  Member:       netto × {SILVER_MULTIPLIER:.0%}  (0 % rabatt)
  Entrepreneur: netto × {GULD_MULTIPLIER:.0%}   (-{int((1-GULD_MULTIPLIER)*100)} % rabatt)
  Krets:        netto × {KRETS_MULTIPLIER:.0%}   (-{int((1-KRETS_MULTIPLIER)*100)} % rabatt, SEK)

  Ladda upp CSVs i SparkLayer → Price Lists
""")

    # Uppdatera Excel-masterfile
    update_pricelist_excel(products, rates)


def update_pricelist_excel(products: list, rates: dict):
    """
    Uppdaterar reCarpet-pricelists-master.xlsx med:
    - Live valutakurser i Exchange Rates-sheeten
    - Alla ORAK-produkter i Entrepreneur- och Members-sheetsen
    - Uppdaterade Export-sheets med rätt antal rader
    """
    if not PRICELIST_EXCEL.exists():
        print(f"  Warning: Excel-masterfile ej hittad: {PRICELIST_EXCEL}")
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  Warning: openpyxl ej installerat — pip3 install openpyxl")
        return

    print(f"  Uppdaterar {PRICELIST_EXCEL.name}...")
    wb = load_workbook(PRICELIST_EXCEL)

    eur_sek = rates["EUR_SEK"]
    eur_nok = rates["EUR_NOK"]
    eur_dkk = rates["EUR_DKK"]

    # ── 1. Exchange Rates-sheet ───────────────────────────────────────────────
    ws_rates = wb["Exchange Rates"]
    # C4=NOK, C5=DKK, C6=EUR (alla = × SEK)
    ws_rates["C4"] = round(eur_nok / eur_sek, 6)  # NOK per SEK
    ws_rates["C5"] = round(eur_dkk / eur_sek, 6)  # DKK per SEK
    ws_rates["C6"] = round(1 / eur_sek, 6)         # EUR per SEK
    # Uppdatera not med källa och datum
    rate_note = (
        f"Live-kurs uppdaterad {rates.get('date','')}: "
        f"1 EUR = {eur_sek} SEK | {eur_nok} NOK | {eur_dkk} DKK"
    )
    ws_rates["A2"] = rate_note

    # ── 2. Bygg produktlista (EUR-priser → SEK/NOK/DKK) ──────────────────────
    product_rows = []   # (sku, sek_member, sek_entre, nok_member, nok_entre, dkk_member, dkk_entre, eur_member, eur_entre)
    for p in products:
        raw_sku = safe(p.get("sku"))
        if not raw_sku or should_skip(raw_sku):
            continue
        sku = make_recarpet_sku(raw_sku)
        try:
            eur_price = float(p.get("price", 0) or 0)
        except (ValueError, TypeError):
            continue
        if eur_price <= 0:
            continue

        def px(rate, mult):
            return round(eur_price * rate * mult, 2)

        product_rows.append({
            "sku":        sku,
            "sek_member": px(eur_sek, SILVER_MULTIPLIER),
            "sek_entre":  px(eur_sek, GULD_MULTIPLIER),
            "sek_krets":  px(eur_sek, KRETS_MULTIPLIER),
            "nok_member": px(eur_nok, SILVER_MULTIPLIER),
            "nok_entre":  px(eur_nok, GULD_MULTIPLIER),
            "dkk_member": px(eur_dkk, SILVER_MULTIPLIER),
            "dkk_entre":  px(eur_dkk, GULD_MULTIPLIER),
            "eur_member": px(1.0,     SILVER_MULTIPLIER),
            "eur_entre":  px(1.0,     GULD_MULTIPLIER),
        })

    # ── 3. Entrepreneur-sheet ─────────────────────────────────────────────────
    _rebuild_price_sheet(wb["Entrepreneur"], product_rows, "sek_entre", "nok_entre", "dkk_entre", "eur_entre")

    # ── 4. Members-sheet ──────────────────────────────────────────────────────
    _rebuild_price_sheet(wb["Members"], product_rows, "sek_member", "nok_member", "dkk_member", "eur_member")

    # ── 5. Export-sheets ──────────────────────────────────────────────────────
    export_map = {
        "Export entrepreneur-sek": ("sek_entre",  "Entrepreneur", "C"),
        "Export entrepreneur-nok": ("nok_entre",  "Entrepreneur", "D"),
        "Export entrepreneur-dkk": ("dkk_entre",  "Entrepreneur", "E"),
        "Export entrepreneur-eur": ("eur_entre",  "Entrepreneur", "F"),
        "Export member-sek":       ("sek_member", "Members",      "C"),
        "Export member-nok":       ("nok_member", "Members",      "D"),
        "Export member-dkk":       ("dkk_member", "Members",      "E"),
        "Export member-eur":       ("eur_member", "Members",      "F"),
    }
    for sheet_name, (price_key, src_sheet, _) in export_map.items():
        _rebuild_export_sheet(wb[sheet_name], product_rows, price_key)

    # ── Bulk Upload-sheet (SKU, PRICE, PRICE_LIST_SLUG) ───────────────────────
    # SparkLayer slugs matchar handles i Price Lists-vyn
    SLUG_MAP = [
        ("sek_entre",  "entrepreneur"),
        ("nok_entre",  "entrepreneur-nok"),
        ("dkk_entre",  "entrepreneur-dkk"),
        ("eur_entre",  "entrepreneur-eur"),
        ("sek_member", "member"),
        ("nok_member", "member-nok"),
        ("dkk_member", "member-dkk"),
        ("eur_member", "member-eur"),
        ("sek_krets",  "krets"),
    ]

    if "Bulk Upload" not in wb.sheetnames:
        wb.create_sheet("Bulk Upload")
    ws_bulk = wb["Bulk Upload"]
    # Rensa och sätt header
    ws_bulk.delete_rows(1, ws_bulk.max_row)
    ws_bulk.append(["SKU", "PRICE", "PRICE_LIST_SLUG"])
    for p in product_rows:
        for price_key, slug in SLUG_MAP:
            ws_bulk.append([p["sku"], p[price_key], slug])

    # Spara bulk-CSV direkt till sparklayer-pricelists/
    bulk_csv_path = SPARKLAYER_DIR / "sparklayer-bulk-upload.csv"
    with open(bulk_csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["SKU", "PRICE", "PRICE_LIST_SLUG"])
        for p in product_rows:
            for price_key, slug in SLUG_MAP:
                writer.writerow([p["sku"], p[price_key], slug])
    print(f"  OK  sparklayer-bulk-upload.csv  ({len(product_rows) * len(SLUG_MAP)} rader, {len(product_rows)} SKUs × {len(SLUG_MAP)} listor)")

    wb.save(PRICELIST_EXCEL)

    # Recalkulera formler via LibreOffice (beräknar NOK/DKK/EUR-kolumnerna)
    recalc_script = BASE_DIR.parent / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"
    if recalc_script.exists():
        import subprocess, json as _json
        try:
            result = subprocess.run(
                ["python3", str(recalc_script), str(PRICELIST_EXCEL), "60"],
                capture_output=True, text=True, timeout=120
            )
            recalc_data = _json.loads(result.stdout)
            if recalc_data.get("status") == "success":
                print(f"  Formler recalkulerade: {recalc_data.get('total_formulas',0)} st")
            else:
                print(f"  Warning: recalc: {recalc_data}")
        except Exception as e:
            print(f"  Warning: recalc misslyckades: {e}")

    print(f"  OK  {PRICELIST_EXCEL.name}  ({len(product_rows)} produkter, kurs: {rates.get('source','?')})")


def _rebuild_price_sheet(ws, product_rows: list, sek_key, nok_key, dkk_key, eur_key):
    """Rensar datarader (rad 5+) och skriver om med aktuella produkter."""
    # Ta bort befintliga datarader
    max_row = ws.max_row
    if max_row >= 5:
        ws.delete_rows(5, max_row - 4)

    for i, p in enumerate(product_rows, start=5):
        ws.cell(i, 1, p["sku"])
        ws.cell(i, 2, 1)
        ws.cell(i, 3, p[sek_key])
        ws.cell(i, 4, f"=ROUND(C{i}*'Exchange Rates'!$C$4,2)")
        ws.cell(i, 5, f"=ROUND(C{i}*'Exchange Rates'!$C$5,2)")
        ws.cell(i, 6, f"=ROUND(C{i}*'Exchange Rates'!$C$6,2)")
        ws.cell(i, 7, f'=A{i}&","&B{i}&","&TEXT(C{i},"0.00")')
        ws.cell(i, 8, f'=A{i}&","&B{i}&","&TEXT(D{i},"0.00")')
        ws.cell(i, 9, f'=A{i}&","&B{i}&","&TEXT(E{i},"0.00")')
        ws.cell(i, 10, f'=A{i}&","&B{i}&","&TEXT(F{i},"0.00")')


def _rebuild_export_sheet(ws, product_rows: list, price_key: str):
    """Rensar datarader (rad 4+) och skriver om med sku, qty, pris."""
    max_row = ws.max_row
    if max_row >= 4:
        ws.delete_rows(4, max_row - 3)

    for i, p in enumerate(product_rows, start=4):
        ws.cell(i, 1, p["sku"])
        ws.cell(i, 2, 1)
        ws.cell(i, 3, p[price_key])


# ─── Main ─────────────────────────────────────────────────────────────────────

PRODUCT_TYPE_MAP = {
    "reemploi":        "Carpet Tile - Reemploi",
    "reemploi":        "Carpet Tile - Reemploi",
    "fin de serie":    "Carpet Tile - Fin de serie",
    "produit a venir": "Carpet Tile - A venir",
}


def normalise_label(s: str) -> str:
    """Remove accents for map lookup."""
    replacements = {"é": "e", "è": "e", "ê": "e", "à": "a", "â": "a", "ô": "o", "û": "u", "î": "i"}
    result = s.lower()
    for k, v in replacements.items():
        result = result.replace(k, v)
    return result


def should_skip(raw_sku: str) -> bool:
    """Returnerar True om raden är en intern ORAK-anteckning och ska hoppas över."""
    if raw_sku in SKIP_RAW_SKUS:
        return True
    raw_lower = raw_sku.lower()
    return any(pat in raw_lower for pat in SKIP_PATTERNS)


# Suffix-nyckelord att strippa (allt från första förekomst av dessa och framåt)
import re as _re
_SUFFIX_KEYWORDS = _re.compile(
    r'[\s_-]+(attente|nettoyage|surplus|pour\s|icade|atelier|welltek|bloqué|bloque|dv\d)',
    flags=_re.IGNORECASE
)

def make_recarpet_sku(raw: str) -> str:
    """
    Omvandlar ORAKs råa identifierare till ett reCarpet-standardiserat SKU.
    Format: RC-ORAK-<rensat-id>
    Exempel:
      GIZ2202010                    → RC-ORAK-GIZ2202010
      2312005-01_attente nettoyage  → RC-ORAK-2312005-01
      240803-01-surplus             → RC-ORAK-240803-01
      251003-01_welltek_12/25       → RC-ORAK-251003-01
      SAPE_2                        → RC-ORAK-SAPE-2
      4700 elegance_atelier_d       → RC-ORAK-4700-elegance
    """
    base = raw.strip()
    # Hitta första förekomst av ett känt suffix-nyckelord och trunkera där
    m = _SUFFIX_KEYWORDS.search(base)
    if m:
        base = base[:m.start()]
    # Ersätt mellanslag och understreck med bindestreck, rensa avslutande skräp
    base = _re.sub(r'[\s_]+', '-', base)
    base = base.strip('-')
    return "RC-ORAK-" + base


def main():
    parser = argparse.ArgumentParser(description="reCarpet - ORAK Product Import")
    parser.add_argument("--csv",             required=True,      help="Path to ORAK CSV file")
    parser.add_argument("--dry-run",         action="store_true", help="Simulate without writing to Shopify")
    parser.add_argument("--pricelists-only", action="store_true", help="Only generate SparkLayer CSVs")
    parser.add_argument("--fix-sparklayer",  action="store_true", help="Publicera alla befintliga Shopify-produkter till SparkLayer")
    args = parser.parse_args()

    csv_path = Path(args.csv)
    if not csv_path.exists():
        print(f"Error: file not found: {csv_path}")
        sys.exit(1)

    if not args.dry_run and not args.pricelists_only:
        if not SHOPIFY_SHOP or not SHOPIFY_TOKEN:
            print("Error: set SHOPIFY_SHOP and SHOPIFY_TOKEN in .env")
            sys.exit(1)

    mode = "DRY-RUN" if args.dry_run else ("PRICELISTS ONLY" if args.pricelists_only else ("FIX SPARKLAYER" if args.fix_sparklayer else "LIVE"))
    print(f"\nreCarpet - ORAK Import [{mode}]")
    print(f"  CSV: {csv_path.name}")
    print(f"  Shopify: {SHOPIFY_SHOP or '(not active)'}\n")

    print("Loading ORAK data...")
    products = load_orak_csv(csv_path)
    print(f"  {len(products)} products loaded\n")

    # Hämta live valutakurser (EUR → SEK/NOK/DKK)
    print("Hämtar valutakurser...")
    rates = fetch_exchange_rates()

    # Always generate pricelists
    generate_pricelists(products, rates)

    if args.pricelists_only:
        return

    sync_log = load_sync_log()
    existing = {}
    location_id = None
    sparklayer_pub_id = None
    collection_id_map = {}

    if not args.dry_run:
        existing          = get_existing_skus()
        location_id       = get_first_location_id()
        sparklayer_pub_id = get_sparklayer_publication_gid()
        collection_id_map = get_collection_id_map()
        if not collection_id_map:
            print("  OBS: Inga collections hittade. Skapa dem i Shopify Admin först.")
        else:
            print(f"  Collections hittade: {list(collection_id_map.keys())}")

    created = updated = skipped = errors = 0

    print("Syncing products to Shopify...")
    for p in products:
        sku       = safe(p.get("sku"))
        title     = safe(p.get("title"))
        brand     = safe(p.get("brand"), "Unknown")
        label     = safe(p.get("product_label"), "")
        dims      = safe(p.get("dimensions"), "")
        image_url = safe(p.get("image_url"), "")
        tech_url  = safe(p.get("technical_sheet_url"), "")

        try:
            eur_price = float(p.get("price", 0) or 0)
            quantity  = int(float(p.get("quantity", 0) or 0))
        except (ValueError, TypeError):
            eur_price, quantity = 0.0, 0

        # Shopify-priset är listpriset i SEK (ingen B2B-rabatt — SparkLayer hanterar det)
        price = round(eur_price * rates["EUR_SEK"], 2)

        if not sku:
            skipped += 1
            continue

        # Hoppa över ORAKs interna anteckningsrader
        if should_skip(sku):
            print(f"  SKIP     [{sku}] (intern ORAK-rad)")
            skipped += 1
            continue

        # Bygg reCarpet-SKU och spara original som metafält
        raw_orak_sku = sku
        sku = make_recarpet_sku(raw_orak_sku)

        product_type = PRODUCT_TYPE_MAP.get(normalise_label(label), "Carpet Tile")

        # Taggar: alla ORAK-produkter får "atervunna-mattor"; "Produit à venir" får dessutom "kommande"
        label_key = normalise_label(label) if label else ""
        extra_tags = ["kommande"] if label_key in KOMMANDE_LABELS else []
        tags = ",".join(filter(None, [
            "b2b", "b2b-only", "orak", "atervunna-mattor",
            brand.lower(),
            label_key.replace(" ", "-") if label_key else "",
        ] + extra_tags))

        metafields = [
            # Spara ORAKs ursprungliga ID för spårbarhet
            {"namespace": "recarpet", "key": "supplier_sku", "value": raw_orak_sku, "type": "single_line_text_field"},
        ]
        if dims:
            metafields.append({"namespace": "recarpet", "key": "dimensions",         "value": dims,     "type": "single_line_text_field"})
        if tech_url:
            metafields.append({"namespace": "recarpet", "key": "technical_sheet_url","value": tech_url, "type": "url"})
        if label:
            metafields.append({"namespace": "recarpet", "key": "product_label",      "value": label,    "type": "single_line_text_field"})

        is_new = sku not in existing

        # Alla ORAK-produkter → "Återbrukade mattor"
        # "Återbrukade mattor" → Shopify auto-genererar handle: aterbrukade-mattor
        collection_id = (
            collection_id_map.get("aterbrukade-mattor")    # ny collection (skapad från "Återbrukade mattor")
            or collection_id_map.get(ATERVUNNA_HANDLE)    # config-variabel
            or collection_id_map.get("återbrukade mattor") # titel-fallback
            or collection_id_map.get("mattor")             # gammalt handle
        )
        kommande_note = " [kommande]" if label_key in KOMMANDE_LABELS else ""

        if args.dry_run:
            action = "CREATE" if is_new else "UPDATE"
            print(f"  [{action}] {sku}  {title}  price:{price:.2f}  qty:{quantity}  → {ATERVUNNA_HANDLE}{kommande_note}")
            skipped += 1
            continue

        try:
            if is_new:
                product     = create_product(title, brand, product_type, tags, price, sku, image_url)
                product_id  = product["id"]
                variant_id  = product["variants"][0]["id"]
                inv_item_id = product["variants"][0]["inventory_item_id"]

                set_inventory(inv_item_id, location_id, quantity)
                set_metafields(product_id, metafields)

                # Publicera till SparkLayer
                if sparklayer_pub_id:
                    publish_to_sparklayer(product_id, sparklayer_pub_id)

                # Lägg till i rätt produktserie
                if collection_id:
                    add_product_to_collection(product_id, collection_id)

                sync_log[sku] = {"product_id": product_id, "variant_id": variant_id, "inventory_item_id": inv_item_id}
                existing[sku] = sync_log[sku]

                pub_ok = "✓ SparkLayer" if sparklayer_pub_id else "– SparkLayer (ej konfigurerad)"
                col_ok = f"✓ Återbrukade mattor (id:{collection_id})" if collection_id else "– collection ej hittad (kolla handle/titel i Shopify)"
                print(f"  CREATED  {sku} - {title}{kommande_note}  [{pub_ok}] [{col_ok}]")
                created += 1
            else:
                info = existing[sku]
                # Full uppdatering — titel, varumärke, typ, taggar, pris, bild
                update_product(
                    info["product_id"], info["variant_id"],
                    title, brand, product_type, tags, price, image_url
                )
                if info.get("inventory_item_id"):
                    set_inventory(info["inventory_item_id"], location_id, quantity)
                set_metafields(info["product_id"], metafields)

                # Publicera till SparkLayer (idempotent)
                if sparklayer_pub_id:
                    publish_to_sparklayer(info["product_id"], sparklayer_pub_id)

                # Säkerställ att produkten finns i "Återbrukade mattor"
                if collection_id:
                    add_product_to_collection(info["product_id"], collection_id)

                print(f"  UPDATED  {sku} - {title}  qty:{quantity}  price:{price:.2f}{kommande_note}")
                updated += 1

        except Exception as e:
            print(f"  ERROR {sku}: {e}")
            errors += 1

    if not args.dry_run:
        save_sync_log(sync_log)

    print(f"""
-------------------------------------
Done [{mode}]
  Created:  {created}
  Updated:  {updated}
  Skipped:  {skipped}
  Errors:   {errors}
-------------------------------------
""")

    # ── --fix-sparklayer: publicera ALLA befintliga produkter till SparkLayer ──
    if args.fix_sparklayer and not args.dry_run:
        if not sparklayer_pub_id:
            print("fix-sparklayer: SparkLayer publication GID saknas — avbryter.")
            return
        print(f"\nfix-sparklayer: Publicerar {len(existing)} produkter till SparkLayer...")
        fixed = fix_errors = 0
        for i, (sku, info) in enumerate(existing.items(), 1):
            try:
                publish_to_sparklayer(info["product_id"], sparklayer_pub_id)
                fixed += 1
                if i % 50 == 0:
                    print(f"  ...{i}/{len(existing)}")
            except Exception as e:
                print(f"  ERROR {sku}: {e}")
                fix_errors += 1
        print(f"  Klart: {fixed} publicerade, {fix_errors} fel.")


if __name__ == "__main__":
    main()
